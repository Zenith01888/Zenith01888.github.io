"""
Web 上位机主程序 — Flask 服务器 + 后台定时读取示波器 (4通道)
"""

import io
import time
import threading
from datetime import datetime

from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import ENABLED_CHANNELS, MAX_HISTORY_POINTS, CHANNEL_NAMES
from config import CHANNEL_CURRENT_MODE, CHANNEL_FACTOR
from oscilloscope import Oscilloscope

# ---------------------------------------------------------------------------
# Flask 应用
# ---------------------------------------------------------------------------
app = Flask(__name__)

scope = Oscilloscope()

# 历史数据缓存
history: list[dict] = []
latest_data: dict = {"timestamp": None, "channels": {}}

# Mutable storage for current range (not in config.py to keep it clean)
_range_store: dict[int, float] = {1: 30, 2: 30, 3: 30, 4: 30}

# ---------------------------------------------------------------------------
# 路由
# ---------------------------------------------------------------------------


@app.route("/")
def index():
    return render_template("index.html", channels=ENABLED_CHANNELS)


@app.route("/api/data")
def api_data():
    return jsonify(latest_data)


@app.route("/api/history")
def api_history():
    return jsonify(history)


@app.route("/api/status")
def api_status():
    return jsonify({
        "connected": scope.is_connected,
        "resource": scope.resource_name,
        "vendor": scope.vendor,
        "idn": scope.idn,
        "channel_names": CHANNEL_NAMES,
    })


@app.route("/api/channel-names", methods=["GET", "POST"])
def api_channel_names():
    """获取或更新通道名称"""
    if request.method == "POST":
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            return jsonify({"ok": False, "error": "expected JSON object"}), 400
        for ch_str, name in data.items():
            ch = int(ch_str)
            if ch in ENABLED_CHANNELS:
                CHANNEL_NAMES[ch] = str(name).strip() or f"通道 {ch}"
    return jsonify(CHANNEL_NAMES)


@app.route("/api/resources")
def api_resources():
    resources = Oscilloscope.list_resources()
    return jsonify({"resources": resources, "current": scope.resource_name})


@app.route("/api/connect", methods=["POST"])
def api_connect():
    data = request.get_json()
    new_resource = data.get("resource", "") if data else ""
    if not new_resource:
        return jsonify({"ok": False, "error": "missing resource parameter"}), 400
    ok = scope.switch_and_connect(new_resource)
    return jsonify({"ok": ok, "resource": scope.resource_name, "connected": scope.is_connected})


@app.route("/api/settings", methods=["GET", "POST"])
def api_settings():
    """读取或应用示波器挡位/探头/时基设置"""
    if not scope.is_connected:
        return jsonify({"ok": False, "error": "oscilloscope not connected"}), 400
    if request.method == "POST":
        data = request.get_json()
        if data:
            try:
                scope.apply_settings(data)
                return jsonify({"ok": True})
            except Exception as e:
                return jsonify({"ok": False, "error": str(e)}), 500
    # GET
    try:
        return jsonify(scope.get_all_settings())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/run", methods=["POST"])
def api_run():
    """启动示波器采集"""
    if not scope.is_connected:
        return jsonify({"ok": False, "error": "oscilloscope not connected"}), 400
    try:
        scope.run()
        return jsonify({"ok": True, "running": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/stop", methods=["POST"])
def api_stop():
    """停止示波器采集"""
    if not scope.is_connected:
        return jsonify({"ok": False, "error": "oscilloscope not connected"}), 400
    try:
        scope.stop()
        return jsonify({"ok": True, "running": False})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/current-config", methods=["GET", "POST"])
def api_current_config():
    """读取或更新每通道电流测量模式、系数与量程"""
    global CHANNEL_CURRENT_MODE, CHANNEL_FACTOR, _range_store
    if request.method == "POST":
        data = request.get_json()
        if data:
            if "mode" in data:
                for ch_str, mode in data["mode"].items():
                    ch = int(ch_str)
                    if ch in ENABLED_CHANNELS:
                        CHANNEL_CURRENT_MODE[ch] = str(mode)
            if "factor" in data:
                for ch_str, factor in data["factor"].items():
                    ch = int(ch_str)
                    if ch in ENABLED_CHANNELS:
                        CHANNEL_FACTOR[ch] = float(factor)
            if "range" in data:
                for ch_str, rng in data["range"].items():
                    ch = int(ch_str)
                    if ch in ENABLED_CHANNELS:
                        _range_store[ch] = float(rng)
    return jsonify({
        "mode": {str(k): v for k, v in CHANNEL_CURRENT_MODE.items()},
        "factor": {str(k): v for k, v in CHANNEL_FACTOR.items()},
        "range": {str(k): v for k, v in _range_store.items()},
    })


@app.route("/api/waveform/<int:channel>")
def api_waveform(channel: int):
    """获取单通道波形数据"""
    if not scope.is_connected:
        return jsonify({"ok": False, "error": "oscilloscope not connected"}), 400
    if channel not in ENABLED_CHANNELS:
        return jsonify({"ok": False, "error": "invalid channel"}), 400
    try:
        wf = scope.read_waveform(channel)
        return jsonify({"ok": True, "channel": channel, "waveform": wf})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# Column format map for Excel export
_COLUMN_FORMATS = {
    0: "0.000000",   # rms
    1: "0.000000",   # vmax
    2: "0.000000",   # vmin
    3: "0.000000",   # vpp
    4: "0.000000",   # voltage (average)
    5: "0.000000",   # current
    6: "0.00",       # frequency
    7: "0.00",       # duty_cycle
    8: "0.000000000", # pulse_width
    9: "0.000000000", # rise_time
    10: "0.000000000", # fall_time
    11: "0.00",      # overshoot
}

def _column_format(offset: int) -> str:
    return _COLUMN_FORMATS.get(offset, "0.000000")

@app.route("/api/export")
def api_export():
    """导出历史数据为 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "示波器数据"

    # 表头
    headers = ["时间"]
    for ch in ENABLED_CHANNELS:
        name = CHANNEL_NAMES.get(ch, f"CH{ch}")
        headers.append(f"{name} RMS(V)")
        headers.append(f"{name} Max(V)")
        headers.append(f"{name} Min(V)")
        headers.append(f"{name} P-P(V)")
        headers.append(f"{name} 平均值(V)")
        headers.append(f"{name} 电流(A)")
        headers.append(f"{name} 频率(Hz)")
        headers.append(f"{name} 占空比(%)")
        headers.append(f"{name} 脉宽(s)")
        headers.append(f"{name} 上升时间(s)")
        headers.append(f"{name} 下降时间(s)")
        headers.append(f"{name} 过冲(%)")

    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="94A3B8"),
        right=Side(style="thin", color="94A3B8"),
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    data_font = Font(name="Consolas", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(history, 2):
        ws.cell(row=row_idx, column=1, value=record.get("timestamp", "")).font = data_font
        ws.cell(row=row_idx, column=1).alignment = data_align
        ws.cell(row=row_idx, column=1).border = thin_border

        chs = record.get("channels", {})
        for ch in ENABLED_CHANNELS:
            ch_data = chs.get(str(ch), {}) if all(isinstance(k, str) for k in chs) else chs.get(ch, {})
            base_col = (ch - 1) * 12 + 2
            vals = [
                ch_data.get("rms"), ch_data.get("vmax"), ch_data.get("vmin"), ch_data.get("vpp"),
                ch_data.get("voltage"), ch_data.get("current"),
                ch_data.get("frequency"), ch_data.get("duty_cycle"),
                ch_data.get("pulse_width"), ch_data.get("rise_time"),
                ch_data.get("fall_time"), ch_data.get("overshoot"),
            ]
            for offset, val in enumerate(vals):
                cell = ws.cell(row=row_idx, column=base_col + offset)
                if val is not None:
                    cell.value = val
                    cell.number_format = _column_format(offset)
                else:
                    cell.value = "--"
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

    # 列宽
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 20
    for ch in ENABLED_CHANNELS:
        base_col = (ch - 1) * 12 + 2
        for offset in range(12):
            letter = get_column_letter(base_col + offset)
            ws.column_dimensions[letter].width = 16

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"oscilloscope_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# 后台测量循环
# ---------------------------------------------------------------------------


def measure_loop(interval: int = 1):
    global latest_data, history

    while True:
        try:
            if not scope.is_connected:
                print("[后台] 示波器未连接, 尝试重连...")
                scope.connect()

            if scope.is_connected:
                channels_data = scope.read_all_channels()

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # 将 channel key 转为 str 以保持 JSON 兼容
                channels_str = {str(k): v for k, v in channels_data.items()}

                record = {"timestamp": now, "channels": channels_str}
                latest_data = record
                history.append(record)

                if len(history) > MAX_HISTORY_POINTS:
                    history.pop(0)

                # 打印摘要 (测量失败时显示 --)
                def _fmt_val(v, digits=4):
                    if isinstance(v, (int, float)):
                        return f"{v:.{digits}f}"
                    return "--"

                summary = "  ".join(
                    f"CH{ch}: {_fmt_val(d['voltage'])}V"
                    + (f" {_fmt_val(d.get('current'))}A" if d.get("current") is not None else "")
                    + (f" {_fmt_val(d.get('duty_cycle'), 2)}%" if d.get("duty_cycle") is not None else "")
                    for ch, d in channels_data.items()
                )
                print(f"[{now}] {summary}")
            else:
                print("[后台] 连接失败, 1秒后重试...")

        except Exception as e:
            print(f"[后台] 读取异常: {e}")

        time.sleep(interval)


# ---------------------------------------------------------------------------
# 启动入口
# ---------------------------------------------------------------------------


def main():
    print("=" * 50)
    print("  示波器 Web 上位机 (多通道)")
    print(f"  采集通道: {ENABLED_CHANNELS}")
    print("=" * 50)

    scope.connect()

    thread = threading.Thread(target=measure_loop, args=(1,), daemon=True)
    thread.start()
    print("[后台] 测量线程已启动 (间隔=1秒)")

    app.run(host="0.0.0.0", port=5000, debug=False)


if __name__ == "__main__":
    main()
